"""Binary spreadsheet ingestion: ``.xlsx`` / ``.xls`` → tabular text.

The compression pipeline is text-only, so binary spreadsheets enter through this
adapter at the SDK boundary. Each sheet is rendered to CSV text, which then flows
through the normal tabular detection → SmartCrusher path like any other table.

Parsers are optional dependencies (``pip install headroom-ai[spreadsheet]``) and
are imported lazily; a missing dependency fails loudly with an actionable
message rather than silently degrading.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path

__all__ = ["load_spreadsheet"]


def _is_blank_row(row: list[object]) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)


def _rows_to_csv(rows: list[list[object]]) -> str:
    """Render rows to CSV text, dropping fully empty trailing rows.

    openpyxl's ``iter_rows`` walks the sheet's *used range*, which routinely
    extends past the last data row (leftover formatting, a cleared cell), so a
    real sheet commonly ends in ``(None, None, ...)`` tuples. Those were written
    out as blank ``,`` rows — contradicting this function's own contract — so the
    trailing empties are dropped here first.

    ``lineterminator="\\n"`` is load-bearing: ``csv.writer`` defaults to
    ``\\r\\n``, and the old ``.strip("\\n")`` removed only the ``\\n``, leaving a
    dangling ``\\r`` on the final line (and ``,\\r`` residue behind each undropped
    empty row). Using ``\\n`` makes the strip clean.
    """
    last = len(rows)
    while last > 0 and _is_blank_row(rows[last - 1]):
        last -= 1
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    for row in rows[:last]:
        writer.writerow(["" if cell is None else cell for cell in row])
    return buf.getvalue().strip("\n")


def _load_xlsx(path: Path) -> dict[str, str]:
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover - openpyxl ships in [dev]; defensive guard
        raise ImportError(
            "Reading .xlsx files requires openpyxl. "
            "Install it with: pip install headroom-ai[spreadsheet]"
        ) from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, str] = {}
    try:
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            text = _rows_to_csv(rows)
            if text.strip():
                sheets[ws.title] = text
    finally:
        wb.close()
    return sheets


def _load_xls(
    path: Path,
) -> dict[str, str]:  # pragma: no cover - legacy .xls; needs optional xlrd + binary fixture
    try:
        import xlrd
    except ImportError as e:
        raise ImportError(
            "Reading legacy .xls files requires xlrd. "
            "Install it with: pip install headroom-ai[spreadsheet]"
        ) from e

    book = xlrd.open_workbook(str(path))
    sheets: dict[str, str] = {}
    for sheet in book.sheets():
        rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        text = _rows_to_csv(rows)
        if text.strip():
            sheets[sheet.name] = text
    return sheets


def load_spreadsheet(path: str | Path) -> dict[str, str]:
    """Load a spreadsheet file into ``{sheet_name: csv_text}``.

    Args:
        path: Path to a ``.xlsx`` or ``.xls`` file.

    Returns:
        Mapping of sheet name to CSV-rendered text (empty sheets omitted).

    Raises:
        FileNotFoundError: If the path does not exist.
        ValueError: If the file extension is unsupported.
        ImportError: If the required parser dependency is not installed.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Spreadsheet not found: {p}")

    suffix = p.suffix.lower()
    if suffix == ".xlsx":
        return _load_xlsx(p)
    if suffix == ".xls":
        return _load_xls(p)  # pragma: no cover - legacy .xls path, see _load_xls
    raise ValueError(f"Unsupported spreadsheet format '{suffix}'. Supported: .xlsx, .xls")
